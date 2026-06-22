"""xlsx/xls文件扫描器 - 递归扫描目录并提取子表名称"""
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from typing import Dict, List, Tuple, Callable
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from python_calamine import CalamineWorkbook
    HAS_CALAMINE = True
except ImportError:
    HAS_CALAMINE = False

from openpyxl import load_workbook
import xlrd

# XML 命名空间
NS = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

# 编译正则：从 xl/workbook.xml 直接提取 sheet name，比 DOM 解析快 2-3x
# 格式: <sheet name="Sheet1" sheetId="1" r:id="rId1"/>
_SHEET_NAME_RE = re.compile(rb'<sheet\s[^>]*?name="([^"]*)"')

class XlsxScanner:
    def __init__(self, max_workers: int = 8, use_calamine: bool = True):
        self.supported_extensions = ['.xlsx', '.xlsm', '.xls']
        self.max_workers = max_workers
        self.use_calamine = use_calamine

    @staticmethod
    def _trim_preview_row(row_data: List[str]) -> List[str]:
        """移除预览行末尾连续空单元格，避免展示虚高的空列。"""
        last_non_empty_index = -1
        for index, value in enumerate(row_data):
            if value != '':
                last_non_empty_index = index

        if last_non_empty_index == -1:
            return []

        return row_data[:last_non_empty_index + 1]

    @staticmethod
    def _cell_matches(value: str, keyword: str, match_mode: str) -> bool:
        """按当前匹配模式判断单元格是否命中关键字。"""
        if not keyword:
            return False

        haystack = value.lower()
        needle = keyword.lower()
        normalized_mode = match_mode or 'fuzzy'

        if normalized_mode == 'exact':
            return haystack == needle
        if normalized_mode == 'prefix':
            return haystack.startswith(needle)
        return needle in haystack

    def is_xlsx_file(self, filepath: str) -> bool:
        """检查是否为支持的表格文件"""
        ext = os.path.splitext(filepath)[1].lower()
        return ext in self.supported_extensions

    @staticmethod
    def _is_xls_format(filepath: str) -> bool:
        """判断是否为旧版 .xls 格式"""
        return os.path.splitext(filepath)[1].lower() == '.xls'

    def get_sheet_names(self, filepath: str) -> List[str]:
        """获取表格文件的所有子表名称"""
        if self._is_xls_format(filepath):
            return self._get_sheet_names_xls(filepath)

        try:
            # xlsx/xlsm 是 zip 文件，用正则直接从 workbook.xml 提取 sheet 名
            # 比 ET.parse 快 2-3x，workbook.xml 结构简单固定无需 DOM
            with zipfile.ZipFile(filepath, 'r') as zf:
                raw = zf.read('xl/workbook.xml')
                names = [name.decode('utf-8') for name in _SHEET_NAME_RE.findall(raw)]
                if names:
                    return names
            return []
        except Exception:
            # 正则失败时回退到 DOM 解析（处理非标准格式），最后尝试 openpyxl
            try:
                with zipfile.ZipFile(filepath, 'r') as zf:
                    with zf.open('xl/workbook.xml') as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        sheets = root.findall('.//main:sheet', NS)
                        if sheets:
                            return [s.get('name', f'Sheet{i+1}') for i, s in enumerate(sheets)]
                return []
            except Exception:
                return self._get_sheet_names_slow(filepath)

    def _get_sheet_names_xls(self, filepath: str) -> List[str]:
        """使用 xlrd 获取 .xls 文件的子表名称"""
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            return wb.sheet_names()
        except Exception as e:
            print(f"警告: 读取 .xls 文件失败 {filepath}: {e}")
            return []

    def _get_sheet_names_slow(self, filepath: str) -> List[str]:
        """备用方式获取子表名称（使用 openpyxl）"""
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return sheet_names
        except Exception as e:
            print(f"警告: 无法读取文件 {filepath}: {e}")
            return []

    def extract_cell_texts(self, filepath: str, sheet_names: List[str],
                           max_chars_per_sheet: int = 20000) -> List[str]:
        """
        提取每个 sheet 的单元格内容（拼接为字符串），用于索引。
        优先使用 calamine（Rust 级解析速度），回退到 openpyxl。
        返回与 sheet_names 平行的字符串列表。
        """
        if self._is_xls_format(filepath):
            return self._extract_cell_texts_xls(filepath, sheet_names, max_chars_per_sheet)

        # 跳过超大文件（>200MB），避免 OOM
        try:
            file_size_mb = os.path.getsize(filepath) / (1024 * 1024)
        except OSError:
            file_size_mb = 0
        if file_size_mb > 80:
            print(f"警告: 跳过超大文件 {os.path.basename(filepath)} ({file_size_mb:.0f}MB)，避免内存溢出")
            return [''] * len(sheet_names)

        if self.use_calamine and HAS_CALAMINE:
            try:
                return self._extract_cell_texts_calamine(
                    filepath, sheet_names, max_chars_per_sheet, file_size_mb
                )
            except Exception as e:
                print(f"calamine 提取失败，回退到 openpyxl: {os.path.basename(filepath)}: {e}")

        return self._extract_cell_texts_openpyxl(filepath, sheet_names, max_chars_per_sheet, file_size_mb)

    def _extract_cell_texts_calamine(self, filepath: str, sheet_names: List[str],
                                      max_chars_per_sheet: int, file_size_mb: float) -> List[str]:
        """calamine 版本：Rust 级 XML 解析，比 openpyxl 快 5-10x"""
        wb = CalamineWorkbook.from_path(filepath)
        results = []
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheet_names:
                results.append('')
                continue
            try:
                sheet = wb.get_sheet_by_name(sheet_name)
            except (KeyError, ValueError):
                results.append('')
                continue
            parts = []
            char_count = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell is not None and cell != '':
                        val = str(cell)
                        parts.append(val)
                        char_count += len(val) + 1
                        if char_count > max_chars_per_sheet:
                            break
                if char_count > max_chars_per_sheet:
                    break
            results.append(' '.join(parts))
        return results

    def _extract_cell_texts_openpyxl(self, filepath: str, sheet_names: List[str],
                                      max_chars_per_sheet: int, file_size_mb: float) -> List[str]:
        """openpyxl 回退版本"""
        wb = None
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            results = []
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    results.append('')
                    continue
                ws = wb[sheet_name]
                parts = []
                char_count = 0
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            val = str(cell.value)
                            parts.append(val)
                            char_count += len(val) + 1
                            if char_count > max_chars_per_sheet:
                                break
                    if char_count > max_chars_per_sheet:
                        break
                results.append(' '.join(parts))
            return results
        except MemoryError:
            print(f"警告: 提取单元格文本内存不足 {os.path.basename(filepath)} ({file_size_mb:.0f}MB)")
            return [''] * len(sheet_names)
        except Exception as e:
            print(f"警告: 提取单元格文本失败 {os.path.basename(filepath)}: {e}")
            return [''] * len(sheet_names)
        finally:
            if wb is not None:
                wb.close()

    def _extract_cell_texts_xls(self, filepath: str, sheet_names: List[str],
                                 max_chars_per_sheet: int = 50000) -> List[str]:
        """xlrd 版本：提取 .xls 文件的单元格内容"""
        wb = None
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            results = []
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheet_names():
                    results.append('')
                    continue
                ws = wb.sheet_by_name(sheet_name)
                parts = []
                char_count = 0
                for row_idx in range(ws.nrows):
                    for col_idx in range(ws.ncols):
                        val = ws.cell_value(row_idx, col_idx)
                        if val is not None and val != '':
                            s = str(val)
                            parts.append(s)
                            char_count += len(s) + 1
                            if char_count > max_chars_per_sheet:
                                break
                    if char_count > max_chars_per_sheet:
                        break
                results.append(' '.join(parts))
            return results
        except MemoryError:
            print(f"警告: 提取 .xls 单元格文本内存不足 {os.path.basename(filepath)}")
            return [''] * len(sheet_names)
        except Exception as e:
            print(f"警告: 提取 .xls 单元格文本失败 {os.path.basename(filepath)}: {e}")
            return [''] * len(sheet_names)
        finally:
            if wb is not None:
                wb.release_resources()

    def _read_sheet_preview_calamine(self, filepath: str, sheet_name: str,
                                      max_rows: int, max_cols: int,
                                      start_row: int, start_col: int) -> List[List[str]]:
        """calamine 版本：读取 sheet 指定窗口数据"""
        wb = CalamineWorkbook.from_path(filepath)
        if sheet_name not in wb.sheet_names:
            return []
        sheet = wb.get_sheet_by_name(sheet_name)
        # calamine 的 rows 是懒加载迭代器，手动构建指定窗口切片
        data = []
        target_row_end = max(start_row, 1) + max_rows - 1
        for row_idx, row in enumerate(sheet.iter_rows()):
            excel_row = row_idx + 1
            if excel_row < max(start_row, 1):
                continue
            if excel_row > target_row_end:
                break
            col_start_0 = max(start_col, 1) - 1
            col_end_0 = col_start_0 + max_cols
            sliced = row[col_start_0:col_end_0]
            row_data = [str(c) if c is not None else '' for c in sliced]
            data.append(self._trim_preview_row(row_data))
        return data

    def read_sheet_preview(self, filepath: str, sheet_name: str,
                           max_rows: int = 20, max_cols: int = 50,
                           start_row: int = 1, start_col: int = 1) -> List[List[str]]:
        """
        读取单个 sheet 指定窗口的数据。优先使用 calamine。
        """
        if self._is_xls_format(filepath):
            return self._read_sheet_preview_xls(filepath, sheet_name, max_rows, max_cols, start_row, start_col)

        if self.use_calamine and HAS_CALAMINE:
            try:
                return self._read_sheet_preview_calamine(
                    filepath, sheet_name, max_rows, max_cols, start_row, start_col
                )
            except BaseException as e:
                print(f"calamine 读取预览失败，回退到 openpyxl: {e}")

        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                wb.close()
                return []
            ws = wb[sheet_name]
            data = []
            for row in ws.iter_rows(
                min_row=max(start_row, 1),
                max_row=max(start_row, 1) + max_rows - 1,
                min_col=max(start_col, 1),
                max_col=max(start_col, 1) + max_cols - 1,
            ):
                row_data = [str(cell.value) if cell.value is not None else '' for cell in row]
                data.append(self._trim_preview_row(row_data))
            wb.close()
            return data
        except Exception as e:
            print(f"警告: 读取预览数据失败 {filepath}: {e}")
            return []

    def _read_sheet_preview_xls(self, filepath: str, sheet_name: str,
                                 max_rows: int = 20, max_cols: int = 50,
                                 start_row: int = 1, start_col: int = 1) -> List[List[str]]:
        """xlrd 版本：读取 .xls 文件的预览数据"""
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            if sheet_name not in wb.sheet_names():
                wb.release_resources()
                return []
            ws = wb.sheet_by_name(sheet_name)
            data = []
            row_start = max(start_row - 1, 0)
            row_end = min(ws.nrows, row_start + max_rows)
            col_start = max(start_col - 1, 0)
            col_end = min(ws.ncols, col_start + max_cols)
            for row_idx in range(row_start, row_end):
                row_data = []
                for col_idx in range(col_start, col_end):
                    val = ws.cell_value(row_idx, col_idx)
                    row_data.append(str(val) if val is not None and val != '' else '')
                data.append(self._trim_preview_row(row_data))
            wb.release_resources()
            return data
        except Exception as e:
            print(f"警告: 读取 .xls 预览数据失败 {filepath}: {e}")
            return []

    # ---- 合并的 read_sheet_with_hits：一次打开文件，返回命中和预览数据 ----

    def read_sheet_with_hits(self, filepath: str, sheet_name: str,
                              keyword: str = None, match_mode: str = 'fuzzy',
                              max_hits: int = 200, preview_rows: int = 20,
                              preview_cols: int = 50,
                              start_row: int = None, start_col: int = None) -> tuple:
        """
        打开文件一次，同时完成命中查找和预览数据读取。
        返回 (hits, preview_data, header_row)。

        - hits: 命中的单元格坐标列表（与 find_sheet_matches 返回一致）
        - preview_data: 预览行数据
        - header_row: 第一行（Excel row 1），用作表头

        若提供 start_row/start_col，直接使用该窗口起点；
        否则若有关键字命中，围绕首个命中居中；否则从 row 2 开始。
        """
        if self._is_xls_format(filepath):
            return self._read_sheet_with_hits_xls(
                filepath, sheet_name, keyword, match_mode, max_hits,
                preview_rows, preview_cols, start_row, start_col
            )

        if self.use_calamine and HAS_CALAMINE:
            try:
                return self._read_sheet_with_hits_calamine(
                    filepath, sheet_name, keyword, match_mode, max_hits,
                    preview_rows, preview_cols, start_row, start_col
                )
            except BaseException as e:
                print(f"calamine 读取失败，回退到 openpyxl: {e}")

        return self._read_sheet_with_hits_openpyxl(
            filepath, sheet_name, keyword, match_mode, max_hits,
            preview_rows, preview_cols, start_row, start_col
        )

    def _read_sheet_with_hits_calamine(self, filepath: str, sheet_name: str,
                                        keyword: str, match_mode: str,
                                        max_hits: int, preview_rows: int,
                                        preview_cols: int,
                                        start_row: int, start_col: int) -> tuple:
        """calamine 版本：一次加载全表，在内存中完成命中查找和预览截取"""
        wb = CalamineWorkbook.from_path(filepath)
        if sheet_name not in wb.sheet_names:
            return [], [], []

        sheet = wb.get_sheet_by_name(sheet_name)

        # 一次性拉取所有行（calamine Rust 解析 → Python 对象）
        all_rows = []
        for row in sheet.iter_rows():
            str_row = [str(c) if c is not None else '' for c in row]
            all_rows.append(str_row)

        return self._process_rows_in_memory(
            all_rows, keyword, match_mode, max_hits,
            preview_rows, preview_cols, start_row, start_col
        )

    def _read_sheet_with_hits_openpyxl(self, filepath: str, sheet_name: str,
                                        keyword: str, match_mode: str,
                                        max_hits: int, preview_rows: int,
                                        preview_cols: int,
                                        start_row: int, start_col: int) -> tuple:
        """openpyxl 版本：回退实现"""
        wb = load_workbook(filepath, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [], [], []

        ws = wb[sheet_name]
        all_rows = []
        for row in ws.iter_rows():
            str_row = [str(cell.value) if cell.value is not None else '' for cell in row]
            all_rows.append(str_row)
        wb.close()

        return self._process_rows_in_memory(
            all_rows, keyword, match_mode, max_hits, preview_rows, preview_cols,
            start_row, start_col
        )

    def _process_rows_in_memory(self, all_rows: list, keyword: str, match_mode: str,
                                 max_hits: int, preview_rows: int,
                                 preview_cols: int,
                                 start_row: int = None, start_col: int = None) -> tuple:
        """在已加载的行数据中查找命中和截取预览，文件无关的纯数据操作。"""
        if not all_rows:
            return [], [], []

        # 表头 = 第一行
        header_row = all_rows[0][:preview_cols]
        header_row = self._trim_preview_row(header_row)

        # 查找命中
        hits = []
        first_hit = None
        if keyword:
            for row_idx, row in enumerate(all_rows):
                if len(hits) >= max_hits:
                    break
                for col_idx, val in enumerate(row):
                    if val and self._cell_matches(val, keyword, match_mode):
                        hits.append({
                            'row': row_idx + 1,
                            'col': col_idx + 1,
                            'value': val,
                        })
                        if first_hit is None:
                            first_hit = (row_idx + 1, col_idx + 1)
                        if len(hits) >= max_hits:
                            break

        # 确定预览窗口起点
        if start_row is not None:
            # 调用方显式指定起点
            preview_start_row = max(start_row, 1)
            preview_start_col = max(start_col or 1, 1)
        elif first_hit:
            # 围绕首个命中居中
            preview_start_row = max(first_hit[0] - 3, 1)
            preview_start_col = max(first_hit[1] - 2, 1)
        else:
            # 默认从 row 2 开始（跳过表头）
            preview_start_row = 2
            preview_start_col = 1

        # 从已加载的行中截取预览数据
        preview_data = []
        row_start_0 = max(preview_start_row - 1, 0)
        row_end_0 = min(len(all_rows), row_start_0 + preview_rows)
        col_start_0 = max(preview_start_col - 1, 0)
        col_end_0 = col_start_0 + preview_cols

        for row_idx in range(row_start_0, row_end_0):
            row = all_rows[row_idx]
            sliced = row[col_start_0:col_end_0]
            preview_data.append(self._trim_preview_row(sliced))

        return hits, preview_data, header_row

    def _read_sheet_with_hits_xls(self, filepath: str, sheet_name: str,
                                   keyword: str, match_mode: str,
                                   max_hits: int, preview_rows: int,
                                   preview_cols: int,
                                   start_row: int, start_col: int) -> tuple:
        """xlrd 版本：合并命中查找和预览读取"""
        try:
            wb = xlrd.open_workbook(filepath, on_demand=True)
            if sheet_name not in wb.sheet_names():
                wb.release_resources()
                return [], [], []

            ws = wb.sheet_by_name(sheet_name)
            all_rows = []
            for row_idx in range(ws.nrows):
                str_row = []
                for col_idx in range(ws.ncols):
                    val = ws.cell_value(row_idx, col_idx)
                    str_row.append(str(val) if val is not None and val != '' else '')
                all_rows.append(str_row)
            wb.release_resources()

            return self._process_rows_in_memory(
                all_rows, keyword, match_mode, max_hits,
                preview_rows, preview_cols, start_row, start_col
            )
        except Exception as e:
            print(f"警告: 读取 .xls 命中/预览失败 {filepath}: {e}")
            return [], [], []

    def find_sheet_matches(self, filepath: str, sheet_name: str, keyword: str,
                           match_mode: str = 'fuzzy', max_hits: int = 200) -> List[Dict]:
        """返回单个 sheet 中命中的单元格坐标和内容（保留向后兼容）。"""
        if not keyword:
            return []

        hits, _, _ = self.read_sheet_with_hits(
            filepath, sheet_name, keyword=keyword, match_mode=match_mode,
            max_hits=max_hits, preview_rows=0, preview_cols=0
        )
        return hits

    def scan_directory_incremental(self, directory: str, index_manager, progress_callback: Callable = None) -> Tuple[int, int, int]:
        """
        增量扫描目录，只更新有变化的文件（多线程并发）
        返回: (新增文件数, 更新文件数, 删除文件数)
        """
        # 1) scandir 走树，收集 (filepath -> mtime)。
        #    用 DirEntry.stat() 避免每次单独 syscall；scandir 本身比 os.walk 快很多（Windows 上尤其明显）。
        all_files = {}
        for root, entry in self._walk_files(directory):
            if not self.is_xlsx_file(entry.name):
                continue
            try:
                stat = entry.stat()
            except OSError:
                continue
            all_files[os.path.join(root, entry.name)] = stat.st_mtime

        # 2) 一次查询拿到当前索引快照，避免 N 次回查
        indexed = index_manager.get_all_files_indexed()

        # 3) 找出需要处理的文件，按目录排序以提升磁盘局部性
        #    同一目录的文件在磁盘上通常相邻，顺序读取可减少 HDD 寻道、改善 SSD 预读命中
        files_to_process = []
        for filepath, mtime in all_files.items():
            existing = indexed.get(filepath)
            if existing is None or existing[1] != mtime:
                files_to_process.append((filepath, os.path.basename(filepath), mtime))
        files_to_process.sort(key=lambda x: x[0])  # 按完整路径排序（等同于按目录+文件排序）

        # 4) 并发提取 sheet 名（进度每 16 个文件上报一次，减少 GUI 信号开销）
        total_files = len(files_to_process)
        if progress_callback:
            progress_callback(0, total_files)

        pending_updates = []
        if files_to_process:
            processed = 0
            _PROGRESS_INTERVAL = 16  # 批量进度间隔
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {
                    executor.submit(self.get_sheet_names, fp): (fp, fn, mt)
                    for fp, fn, mt in files_to_process
                }

                for future in as_completed(futures):
                    fp, fn, mt = futures[future]
                    processed += 1
                    try:
                        sheet_names = future.result()
                        if sheet_names:
                            pending_updates.append((fn, fp, mt, sheet_names))
                    except Exception as e:
                        print(f"警告: 处理文件失败 {fp}: {e}")
                    finally:
                        if progress_callback and (
                            processed % _PROGRESS_INTERVAL == 0 or processed == total_files
                        ):
                            progress_callback(processed, total_files)

        # 5) 单事务批量写库（一次 commit、一次 fsync）
        added, updated = index_manager.upsert_files_batch(pending_updates, indexed)

        # 6) 索引中存在但磁盘上已消失的文件 → 按 id 批量删除
        files_on_disk = set(all_files.keys())
        file_ids_to_delete = [fid for fp, (fid, _) in indexed.items() if fp not in files_on_disk]
        deleted = index_manager.delete_files_by_ids(file_ids_to_delete)

        return (added, updated, deleted)

    @staticmethod
    def _walk_files(directory: str):
        """DFS 遍历，跳过隐藏目录，yield (parent_dir, DirEntry)。

        替代 os.walk：scandir 在 Windows 上明显更快，且 DirEntry 已经缓存了
        findfirst/findnext 的结果，后续 entry.stat() 不需要再发一次 syscall。
        """
        stack = [directory]
        while stack:
            root = stack.pop()
            try:
                with os.scandir(root) as it:
                    subdirs = []
                    for entry in it:
                        try:
                            if entry.is_dir(follow_symlinks=False):
                                if not entry.name.startswith('.'):
                                    subdirs.append(entry.path)
                            else:
                                yield root, entry
                        except OSError:
                            continue
                    for sub in reversed(subdirs):
                        stack.append(sub)
            except (PermissionError, FileNotFoundError, NotADirectoryError, OSError):
                continue
