"""XlsxSearcher 三项优化的正确性 + 性能对比测试。

用法： python benchmarks/bench.py

覆盖：
  1. 预览流式/窗口化读取 —— 与"全量载入再切片"的旧实现做结果一致性断言 + 耗时对比
  2. FTS5 单元格搜索 —— 与 LIKE '%kw%' 旧实现做结果一致性断言 + 耗时对比
  3. 连接复用 + upsert 跳过未变 sheets —— 二次扫描耗时 + cell_text 保留断言

不依赖 PyQt，只用到 core/ 模块。使用独立临时 db，不触碰用户索引库。
"""
import os
import sys
import shutil
import sqlite3
import tempfile
import time
import tracemalloc
from collections import deque

# 让脚本能 import 项目根的 core/
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from openpyxl import Workbook  # noqa: E402

from core.indexer import IndexManager  # noqa: E402
from core.scanner import XlsxScanner  # noqa: E402

# ---------------------------------------------------------------------------
# 临时数据生成
# ---------------------------------------------------------------------------

SHEET_TOKENS = ["config", "player", "level", "weapon", "ui_text", "monster", "shop", "quest"]


def _fill_sheet(ws, rows, cols, seed_text):
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            # 让内容可被关键字命中，且每行不同
            ws.cell(row=r, column=c, value=f"{seed_text}_{r}_{c}_{SHEET_TOKENS[(r + c) % len(SHEET_TOKENS)]}")


def make_medium_file(path, file_idx):
    """50 个中等文件：每个 5 sheet × 200 行 × 20 列。"""
    wb = Workbook()
    first = wb.active
    first.title = "Sheet1"
    _fill_sheet(first, 200, 20, f"f{file_idx}")
    for s in range(2, 6):
        ws = wb.create_sheet(f"Sheet{s}")
        _fill_sheet(ws, 200, 20, f"f{file_idx}s{s}")
    wb.save(path)


def make_large_file(path):
    """1 张大表：1 sheet × 20000 行 × 30 列，并在指定位置埋入 marker。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BigSheet"
    marker_row, marker_col = 15000, 5
    for r in range(1, 20001):
        for c in range(1, 31):
            if r == marker_row and c == marker_col:
                ws.cell(row=r, column=c, value="NEEDLE_CONFIG_MARKER")
            else:
                ws.cell(row=r, column=c, value=f"cell_{r}_{c}_{SHEET_TOKENS[(r + c) % len(SHEET_TOKENS)]}")
    wb.save(path)
    return marker_row, marker_col


# ---------------------------------------------------------------------------
# 旧实现参考（已被优化取代，这里复刻用于一致性对比）
# ---------------------------------------------------------------------------

def old_read_sheet_with_hits(scanner, filepath, sheet_name, keyword=None,
                              match_mode='fuzzy', max_hits=200, preview_rows=20,
                              preview_cols=50, start_row=None, start_col=None):
    """复刻旧实现：一次性把整张表载入内存，再切片。仅用于 benchmark 对比。"""
    from openpyxl import load_workbook
    if scanner._is_xls_format(filepath):
        raise RuntimeError("benchmark 不覆盖 .xls")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return [], [], []
        ws = wb[sheet_name]
        all_rows = []
        for row in ws.iter_rows():
            all_rows.append([str(cell.value) if cell.value is not None else '' for cell in row])
    finally:
        wb.close()

    if not all_rows:
        return [], [], []
    header_row = scanner._trim_preview_row(all_rows[0][:preview_cols])

    hits = []
    first_hit = None
    if keyword:
        for row_idx, row in enumerate(all_rows):
            if len(hits) >= max_hits:
                break
            for col_idx, val in enumerate(row):
                if val and scanner._cell_matches(val, keyword, match_mode):
                    hits.append({'row': row_idx + 1, 'col': col_idx + 1, 'value': val})
                    if first_hit is None:
                        first_hit = (row_idx + 1, col_idx + 1)
                    if len(hits) >= max_hits:
                        break

    if start_row is not None:
        psr = max(start_row, 1)
        psc = max(start_col or 1, 1)
    elif first_hit:
        psr = max(first_hit[0] - 3, 1)
        psc = max(first_hit[1] - 2, 1)
    else:
        psr = 2
        psc = 1

    preview_data = []
    rs0 = max(psr - 1, 0)
    re0 = min(len(all_rows), rs0 + preview_rows)
    cs0 = max(psc - 1, 0)
    ce0 = cs0 + preview_cols
    for ri in range(rs0, re0):
        preview_data.append(scanner._trim_preview_row(all_rows[ri][cs0:ce0]))
    return hits, preview_data, header_row


def old_like_cell_search(db_path, keyword, match_mode='fuzzy'):
    """复刻旧实现的单元格搜索：LIKE '%kw%' 全表扫描。仅用于 benchmark 对比。"""
    conn = sqlite3.connect(db_path)
    try:
        if match_mode == 'exact':
            clause = "LOWER(s.cell_text) = LOWER(?)"
            param = keyword
        elif match_mode == 'prefix':
            clause = "s.cell_text LIKE ? COLLATE NOCASE"
            param = f"{keyword}%"
        else:
            clause = "s.cell_text LIKE ? COLLATE NOCASE"
            param = f"%{keyword}%"
        sql = (
            "SELECT DISTINCT f.filename, f.filepath, s.sheet_name "
            "FROM xlsx_files f LEFT JOIN sheets s ON f.id = s.file_id "
            f"WHERE {clause} ORDER BY LOWER(f.filename), LOWER(s.sheet_name)"
        )
        rows = conn.execute(sql, (param,)).fetchall()
    finally:
        conn.close()
    grouped = {}
    for fn, fp, sn in rows:
        grouped.setdefault(fp, {'filename': fn, 'filepath': fp, 'sheet_names': []})
        if sn:
            grouped[fp]['sheet_names'].append(sn)
    return list(grouped.values())


# ---------------------------------------------------------------------------
# 工具
# ---------------------------------------------------------------------------

def fmt_time(seconds):
    if seconds >= 1:
        return f"{seconds:.2f}s"
    return f"{seconds * 1000:.1f}ms"


def result_set(results):
    """把搜索结果归一成 (filepath, sheet_name) 集合，便于比较。"""
    s = set()
    for r in results:
        for sn in r.get('sheet_names', []):
            s.add((r['filepath'], sn))
    return s


def time_fn(fn, repeat=5):
    best = float('inf')
    for _ in range(repeat):
        t0 = time.perf_counter()
        fn()
        best = min(best, time.perf_counter() - t0)
    return best


def assert_true(cond, msg):
    if not cond:
        raise AssertionError(msg)
    print(f"  [OK] {msg}")


# ---------------------------------------------------------------------------
# 测试主体
# ---------------------------------------------------------------------------

def run():
    tmp = tempfile.mkdtemp(prefix="xlsxsearcher_bench_")
    db_path = os.path.join(tmp, "index.db")
    try:
        print("=" * 70)
        print("生成测试数据...")
        medium_files = []
        for i in range(50):
            p = os.path.join(tmp, f"medium_{i:03d}.xlsx")
            make_medium_file(p, i)
            medium_files.append(p)
        big_path = os.path.join(tmp, "big.xlsx")
        marker_row, marker_col = make_large_file(big_path)
        print(f"  50 个中等文件 + 1 张大表(20000×30)，marker @ R{marker_row}C{marker_col}")

        index_manager = IndexManager(db_path=db_path)
        scanner = XlsxScanner()

        # ---- 扫描 + 深度索引 ----
        print("\n扫描目录（提取 sheet 名）...")
        t0 = time.perf_counter()
        added, updated, deleted = scanner.scan_directory_incremental(tmp, index_manager)
        t_scan1 = time.perf_counter() - t0
        print(f"  首次扫描: +{added} ~{updated} -{deleted}  耗时 {fmt_time(t_scan1)}")

        print("深度索引（提取单元格内容）...")
        t0 = time.perf_counter()
        _run_deep_index(index_manager, scanner)
        t_deep = time.perf_counter() - t0
        stats = index_manager.get_index_status()
        print(f"  深度索引完成: {stats['indexed_cell_sheet_count']} sheets 已索引，耗时 {fmt_time(t_deep)}")

        # =========================================================
        # 测试 1：预览流式/窗口化读取 —— 正确性 + 性能
        # =========================================================
        print("\n" + "=" * 70)
        print("测试 1：预览读取（流式窗口 vs 旧全量载入）")
        keyword = "NEEDLE_CONFIG_MARKER"

        # 1a. 命中查找一致性（hit-determined 窗口）
        new_hits, new_prev, new_header = scanner.read_sheet_with_hits(
            big_path, "BigSheet", keyword=keyword, match_mode='exact')
        old_hits, old_prev, old_header = old_read_sheet_with_hits(
            scanner, big_path, "BigSheet", keyword=keyword, match_mode='exact')
        assert_true(new_hits == old_hits, "关键字命中列表与旧实现一致")
        assert_true(new_prev == old_prev, "命中居中预览窗口与旧实现一致")
        assert_true(new_header == old_header, "表头与旧实现一致")
        assert_true(any(h['row'] == marker_row and h['col'] == marker_col for h in new_hits),
                    "marker 命中坐标正确")

        # 1b. 显式窗口（导航跳转）一致性
        new_h2, new_p2, _ = scanner.read_sheet_with_hits(
            big_path, "BigSheet", keyword=None, start_row=14998, start_col=3)
        old_h2, old_p2, _ = old_read_sheet_with_hits(
            scanner, big_path, "BigSheet", keyword=None, start_row=14998, start_col=3)
        assert_true(new_p2 == old_p2, "显式窗口预览与旧实现一致")
        assert_true(len(new_p2) == 20, "窗口行数 = preview_rows")

        # 1c. 无关键字默认窗口（row 2 起）一致性
        new_h3, new_p3, _ = scanner.read_sheet_with_hits(
            big_path, "BigSheet", keyword=None)
        old_h3, old_p3, _ = old_read_sheet_with_hits(
            scanner, big_path, "BigSheet", keyword=None)
        assert_true(new_p3 == old_p3, "默认窗口(row 2 起)与旧实现一致")

        # 1d. 性能 + 内存：流式 vs 全量
        print("\n  性能对比（大表 20000×30，5 次取最优）:")
        t_new = time_fn(lambda: scanner.read_sheet_with_hits(
            big_path, "BigSheet", keyword=None, start_row=14998, start_col=1))
        t_old = time_fn(lambda: old_read_sheet_with_hits(
            scanner, big_path, "BigSheet", keyword=None, start_row=14998, start_col=1))

        tracemalloc.start()
        scanner.read_sheet_with_hits(big_path, "BigSheet", keyword=None, start_row=14998, start_col=1)
        _, peak_new = tracemalloc.get_traced_memory()
        tracemalloc.reset_peak()
        old_read_sheet_with_hits(scanner, big_path, "BigSheet", keyword=None, start_row=14998, start_col=1)
        _, peak_old = tracemalloc.get_traced_memory()
        tracemalloc.stop()

        print(f"    新(流式窗口): {fmt_time(t_new):>9}   峰值内存 {peak_new/1024/1024:.1f}MB")
        print(f"    旧(全量载入): {fmt_time(t_old):>9}   峰值内存 {peak_old/1024/1024:.1f}MB")
        print(f"    提速 {t_old/t_new:.1f}x，内存降至 {peak_new/peak_old*100:.1f}%")

        # =========================================================
        # 测试 2：FTS5 单元格搜索 —— 正确性 + 性能
        # =========================================================
        print("\n" + "=" * 70)
        print("测试 2：单元格搜索（FTS5 trigram vs LIKE '%kw%'）")
        for kw in ["config", "player", "weapon", "NEEDLE_CONFIG_MARKER"]:
            fts_res = index_manager.search(cell_keyword=kw, match_mode='fuzzy')
            like_res = old_like_cell_search(db_path, kw, 'fuzzy')
            assert_true(result_set(fts_res) == result_set(like_res),
                        f"关键字 '{kw}' FTS 与 LIKE 结果集一致 ({len(result_set(fts_res))} 项)")

        # 短关键字回退 LIKE（<3 字符）
        short_res = index_manager.search(cell_keyword="ui", match_mode='fuzzy')
        short_like = old_like_cell_search(db_path, "ui", 'fuzzy')
        assert_true(result_set(short_res) == result_set(short_like),
                    "短关键字 'ui' 回退 LIKE，结果与 LIKE 一致")

        # 性能：FTS vs LIKE（多次取最优）
        kw = "config"
        t_fts = time_fn(lambda: index_manager.search(cell_keyword=kw, match_mode='fuzzy'), repeat=8)
        t_like = time_fn(lambda: old_like_cell_search(db_path, kw, 'fuzzy'), repeat=8)
        print(f"\n  性能对比（关键字 '{kw}'，8 次取最优）:")
        print(f"    新(FTS5):  {fmt_time(t_fts):>9}")
        print(f"    旧(LIKE):  {fmt_time(t_like):>9}")
        print(f"    提速 {t_like/t_fts:.1f}x")

        # prefix / exact 模式一致性
        for mode, kw in [('prefix', 'conf'), ('exact', 'config')]:
            fts_res = index_manager.search(cell_keyword=kw, match_mode=mode)
            like_res = old_like_cell_search(db_path, kw, mode)
            assert_true(result_set(fts_res) == result_set(like_res),
                        f"模式 {mode} '{kw}' FTS+后过滤 与 LIKE 结果一致")

        # =========================================================
        # 测试 3：upsert 跳过未变 sheets + cell_text 保留
        # =========================================================
        print("\n" + "=" * 70)
        print("测试 3：连接复用 + upsert 跳过未变 sheets")
        # 触碰文件 mtime（内容不变）→ 二次扫描应跳过 sheets 重建
        now = time.time()
        for p in medium_files[:10]:
            os.utime(p, (now, now))
        t0 = time.perf_counter()
        added2, updated2, deleted2 = scanner.scan_directory_incremental(tmp, index_manager)
        t_scan2 = time.perf_counter() - t0
        print(f"  二次扫描(10 文件 mtime 变、内容不变): +{added2} ~{updated2} -{deleted2}  耗时 {fmt_time(t_scan2)}")

        # cell_text 应全部保留（未触发重建）
        pending = index_manager.get_sheets_without_cell_text()
        assert_true(len(pending) == 0, "二次扫描后无 sheet 丢失 cell_text（跳过重建生效）")

        # 真正改动一个文件的 sheet 名 → 应触发该文件重建，其余跳过
        p0 = medium_files[0]
        wb = Workbook()
        ws = wb.active
        ws.title = "RenamedSheet"
        _fill_sheet(ws, 200, 20, "changed")
        wb.save(p0)
        os.utime(p0, (now, now + 5))
        t0 = time.perf_counter()
        a3, u3, d3 = scanner.scan_directory_incremental(tmp, index_manager)
        t_scan3 = time.perf_counter() - t0
        print(f"  三次扫描(1 文件 sheet 名改变): +{a3} ~{u3} -{d3}  耗时 {fmt_time(t_scan3)}")
        # 改动的文件 cell_text 被清空待重新深度索引，其余仍保留
        pending = index_manager.get_sheets_without_cell_text()
        assert_true(len(pending) == 1, f"仅被改动文件的 1 个 sheet 待重新索引（实际 {len(pending)}）")

        # 连接复用：连续多次查询不应报错（线程内长连接）
        for _ in range(50):
            index_manager.search(cell_keyword="config", match_mode='fuzzy')
        assert_true(True, "50 次连续查询（长连接复用）无异常")

        # =========================================================
        # 汇总
        # =========================================================
        print("\n" + "=" * 70)
        print("汇总")
        print(f"  预览读取:  {fmt_time(t_old)} → {fmt_time(t_new)}  ({t_old/t_new:.1f}x 提速)")
        print(f"  单元格搜索: {fmt_time(t_like)} → {fmt_time(t_fts)}  ({t_like/t_fts:.1f}x 提速)")
        print(f"  二次扫描跳过: 首次 {fmt_time(t_scan1)} → 增量 {fmt_time(t_scan2)}")
        print("\n全部正确性断言通过 ✅")
        return 0
    finally:
        # 清理临时目录与 db
        try:
            index_manager.close_thread_connection()
        except Exception:
            pass
        shutil.rmtree(tmp, ignore_errors=True)


def _run_deep_index(index_manager, scanner):
    """复刻 DeepIndexWorker 的核心逻辑（无 GUI）。"""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    from collections import defaultdict
    pending = index_manager.get_sheets_without_cell_text()
    if not pending:
        return
    by_file = defaultdict(list)
    for entry in pending:
        by_file[entry['filepath']].append(entry)

    def _extract(fp, names, ids):
        s = XlsxScanner()
        texts = s.extract_cell_texts(fp, names)
        return [(t, sid) for sid, t in zip(ids, texts)]

    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = {}
        for fp, entries in by_file.items():
            futs[ex.submit(_extract, fp,
                           [e['sheet_name'] for e in entries],
                           [e['sheet_id'] for e in entries])] = fp
        for f in as_completed(futs):
            updates = f.result()
            if updates:
                index_manager.update_sheet_cell_texts_batch(updates)


if __name__ == '__main__':
    sys.exit(run())
