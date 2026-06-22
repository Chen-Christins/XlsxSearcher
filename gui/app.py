"""XlsxSearcher 主界面 - PyQt5 版本"""
import csv
import json
import logging
import os
import sys
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTreeWidget, QTreeWidgetItem, QLabel, QLineEdit, QPushButton,
    QStatusBar, QProgressBar, QMessageBox, QFileDialog, QComboBox,
    QSplitter, QTableWidget, QTableWidgetItem, QHeaderView
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QSettings, QTimer
from PyQt5.QtGui import QBrush, QColor, QIcon

from core.indexer import IndexManager
from core.alias_parser import parse_sheet_alias_file
from core.scanner import XlsxScanner
from core.searcher import Searcher
from utils.file_utils import open_file, open_in_explorer, copy_to_clipboard

LOG_DIR = os.path.join(os.path.expanduser("~"), ".local", "XlsxSearcher")
os.makedirs(LOG_DIR, exist_ok=True)
LOG_PATH = os.path.join(LOG_DIR, "app.log")
logging.basicConfig(
    filename=LOG_PATH,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(threadName)s %(message)s",
)


def _msgbox(parent, level, title, text, buttons=QMessageBox.Ok, default_button=None):
    """统一的 QMessageBox 封装，使用应用图标。

    level: 'info' | 'warn' | 'error' | 'question'（保留参数以便兼容）
    返回: 与 QMessageBox 静态方法一致的按钮常量。
    """
    app_icon = QApplication.windowIcon()
    msg = QMessageBox(parent)
    if not app_icon.isNull():
        msg.setIconPixmap(app_icon.pixmap(64, 64))
    msg.setWindowTitle(title)
    msg.setText(text)
    msg.setStandardButtons(buttons)
    if default_button is not None:
        msg.setDefaultButton(default_button)
    return msg.exec_()


class ScanWorker(QThread):
    """扫描工作线程"""
    finished = pyqtSignal(int, int, int, float)  # added, updated, deleted, duration
    error = pyqtSignal(str)
    progress = pyqtSignal(int, int)  # current, total

    def __init__(self, directory, scanner, index_manager):
        super().__init__()
        self.directory = directory
        self.scanner = scanner
        self.index_manager = index_manager

    def _on_progress(self, current, total):
        self.progress.emit(current, total)

    def run(self):
        import time
        start_time = time.time()
        try:
            added, updated, deleted = self.scanner.scan_directory_incremental(
                self.directory, self.index_manager, progress_callback=self._on_progress
            )
            duration = time.time() - start_time
            self.finished.emit(added, updated, deleted, duration)
        except Exception as e:
            self.error.emit(str(e))


class DeepIndexWorker(QThread):
    """深度索引工作线程 — 并行提取所有未索引 sheet 的单元格内容"""
    finished = pyqtSignal(int, int, float)
    error = pyqtSignal(str)
    progress = pyqtSignal(int, int)

    def __init__(self, index_manager, scanner):
        super().__init__()
        self.index_manager = index_manager
        self.scanner = scanner
        self.warning_summary = ""

    def run(self):
        import time
        from collections import defaultdict
        from multiprocessing import Pipe, get_context
        from core.deep_index_worker import extract_file_cell_texts

        start = time.time()
        processed = 0
        timeout_seconds = 180
        try:
            pending = self.index_manager.get_sheets_without_cell_text()
            total = len(pending)
            if total == 0:
                self.finished.emit(0, 0, 0.0)
                return

            by_file = defaultdict(list)
            for entry in pending:
                by_file[entry['filepath']].append(entry)

            errors = []
            ctx = get_context("spawn")
            for filepath, entries in by_file.items():
                count = len(entries)
                sheet_names = [e['sheet_name'] for e in entries]
                sheet_ids = [e['sheet_id'] for e in entries]
                logging.info("Deep indexing subprocess started: %s (%s sheets)", filepath, count)

                parent_conn, child_conn = Pipe(duplex=False)
                process = ctx.Process(
                    target=extract_file_cell_texts,
                    args=(filepath, sheet_names, child_conn, False),
                )
                try:
                    process.start()
                    child_conn.close()

                    if parent_conn.poll(timeout_seconds):
                        result = parent_conn.recv()
                    else:
                        process.terminate()
                        process.join(5)
                        if process.is_alive():
                            process.kill()
                            process.join(5)
                        result = {
                            "ok": False,
                            "error": f"Timed out after {timeout_seconds} seconds",
                        }

                    process.join(5)
                    if process.is_alive():
                        process.terminate()
                        process.join(5)

                    if process.exitcode not in (0, None) and result.get("ok"):
                        result = {
                            "ok": False,
                            "error": f"Child process exited with code {process.exitcode}",
                        }

                    if result.get("ok"):
                        texts = result.get("texts") or []
                        updates = [(text, sheet_id) for sheet_id, text in zip(sheet_ids, texts)]
                        if updates:
                            self.index_manager.update_sheet_cell_texts_batch(updates)
                        logging.info("Deep indexing subprocess finished: %s", filepath)
                    else:
                        error_text = result.get("error", "unknown child process error")
                        logging.error(
                            "Deep indexing subprocess failed: %s: %s\n%s",
                            filepath,
                            error_text,
                            result.get("traceback", ""),
                        )
                        errors.append(f"{os.path.basename(filepath)}: {error_text}")
                except BaseException as e:
                    logging.exception("Deep indexing subprocess orchestration failed: %s", filepath)
                    errors.append(f"{os.path.basename(filepath)}: {type(e).__name__}: {e}")
                    if process.is_alive():
                        process.terminate()
                        process.join(5)
                finally:
                    parent_conn.close()
                    if process.is_alive():
                        process.kill()
                        process.join(5)
                    processed += count
                    self.progress.emit(processed, total)

            if errors:
                summary = f"{len(errors)} files failed during deep indexing.\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    summary += f"\n...and {len(errors) - 5} more."
                self.warning_summary = summary
                logging.warning(summary)

            self.finished.emit(processed, total, time.time() - start)
        except BaseException as e:
            logging.exception("Deep indexing worker failed")
            self.error.emit(str(e))

class SearchWorker(QThread):
    """搜索工作线程 — 在后台执行 SQLite 查询和结果分组，避免阻塞 UI"""
    finished = pyqtSignal(list)
    error = pyqtSignal(str)

    def __init__(self, searcher, index_manager, sheet_keyword, filename_keyword,
                 cell_keyword, match_mode):
        super().__init__()
        self.searcher = searcher
        self.index_manager = index_manager
        self.sheet_keyword = sheet_keyword
        self.filename_keyword = filename_keyword
        self.cell_keyword = cell_keyword
        self.match_mode = match_mode
        self._cancelled = False

    def cancel(self):
        """标记取消，run() 完成后不发射结果"""
        self._cancelled = True

    def run(self):
        try:
            if not self.sheet_keyword and not self.filename_keyword and not self.cell_keyword:
                results = self.index_manager.get_all_files_with_sheets()
            else:
                results = self.searcher.search(
                    self.sheet_keyword, self.filename_keyword,
                    self.cell_keyword, self.match_mode
                )
            if not self._cancelled:
                self.finished.emit(results)
        except Exception as e:
            if not self._cancelled:
                self.error.emit(str(e))


class XlsxSearcherApp(QMainWindow):
    MAX_SEARCH_HISTORY = 15

    def __init__(self):
        super().__init__()

        # 核心组件
        self.index_manager = IndexManager()
        self.scanner = XlsxScanner(use_calamine=False)
        self.searcher = Searcher(self.index_manager)
        self.settings = QSettings('XlsxSearcher', 'XlsxSearcher')

        # 状态变量
        self.scan_directory = None
        self.search_results = []
        self.search_history = []
        self.is_scanning = False
        self.search_worker = None
        self.current_sort_mode = 'filename_asc'
        self.current_view_mode = 'grouped'
        self._pending_status_prefix = ''
        self.preview_state = {
            'filepath': '',
            'sheet_name': '',
            'start_row': 1,
            'start_col': 1,
            'hits': [],
            'current_hit_index': -1,
            'active_keyword': '',
        }

        self._init_ui()
        self._restore_ui_preferences()
        self._restore_scan_directory()
        self._restore_search_history()
        self._refresh_index_status_label()
        self._check_existing_index()

    def _init_ui(self):
        """初始化UI"""
        # 窗口设置
        self.setWindowTitle("XlsxSearcher - Excel子表搜索工具")
        self.setMinimumSize(1000, 700)
        self.resize(1000, 700)

        # 搜索防抖定时器
        self.search_timer = QTimer()
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self._do_search)

        # macOS 统一标题栏样式
        if sys.platform == 'darwin':
            self.setUnifiedTitleAndToolBarOnMac(True)

        icon_path = _get_icon_path()
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        # 中心部件
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(10, 0, 10, 0)
        main_layout.setSpacing(4)

        # 顶部搜索区域
        top_widget = QWidget()
        top_layout = QHBoxLayout(top_widget)
        if sys.platform == 'darwin':
            top_widget.installEventFilter(self)  # 用于 macOS 窗口拖拽
            top_layout.setContentsMargins(80, 5, 10, 5)
        else:
            top_layout.setContentsMargins(10, 5, 10, 5)
        main_layout.addWidget(top_widget)

        # 目录选择和显示
        self.dir_label = QLabel("未选择目录")
        self.dir_label.setStyleSheet("color: gray")
        self.dir_label.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        top_layout.addWidget(QLabel("扫描目录:"))
        top_layout.addWidget(self.dir_label)
        top_layout.addWidget(QPushButton("选择目录", clicked=self._select_directory))
        top_layout.addWidget(QPushButton("重新扫描", clicked=self._rescan))
        top_layout.addWidget(QPushButton("清空索引", clicked=self._clear_index))
        self.btn_deep_index = QPushButton("深度索引")
        self.btn_deep_index.clicked.connect(self._start_deep_index)
        self.btn_deep_index.setToolTip("提取所有文件的单元格内容以支持单元格搜索")
        top_layout.addWidget(self.btn_deep_index)
        self.btn_import_aliases = QPushButton("导入映射")
        self.btn_import_aliases.clicked.connect(self._import_sheet_aliases)
        self.btn_import_aliases.setToolTip("导入 bat/txt 映射脚本，支持英文名搜索中文子表")
        top_layout.addWidget(self.btn_import_aliases)

        # 搜索区域（两行）
        search_widget = QWidget()
        search_outer = QVBoxLayout(search_widget)
        search_outer.setContentsMargins(10, 0, 10, 10)
        search_outer.setSpacing(4)
        main_layout.addWidget(search_widget)

        # Row 1: 搜索输入（stretch 全给输入框，label 紧贴）
        row1 = QHBoxLayout()
        row1.setSpacing(6)
        self.sheet_entry = QLineEdit()
        self.sheet_entry.setPlaceholderText("子表名称或英文配置名")
        self.sheet_entry.textChanged.connect(self._schedule_search)
        self.sheet_entry.returnPressed.connect(self._on_search_committed)
        self.sheet_entry.editingFinished.connect(self._on_search_committed)
        row1.addWidget(QLabel("子表名称:"))
        row1.addWidget(self.sheet_entry, 1)

        self.filename_entry = QLineEdit()
        self.filename_entry.setPlaceholderText("文件名")
        self.filename_entry.textChanged.connect(self._schedule_search)
        self.filename_entry.returnPressed.connect(self._on_search_committed)
        self.filename_entry.editingFinished.connect(self._on_search_committed)
        row1.addWidget(QLabel("文件名:"))
        row1.addWidget(self.filename_entry, 1)

        self.cell_entry = QLineEdit()
        self.cell_entry.setPlaceholderText("单元格内容")
        self.cell_entry.textChanged.connect(self._schedule_search)
        self.cell_entry.returnPressed.connect(self._on_search_committed)
        self.cell_entry.editingFinished.connect(self._on_search_committed)
        row1.addWidget(QLabel("单元格:"))
        row1.addWidget(self.cell_entry, 1)

        search_outer.addLayout(row1)

        # Row 2: 匹配/排序/视图/历史（stretch 全给下拉框）
        row2 = QHBoxLayout()
        row2.setSpacing(6)
        self.match_mode_combo = QComboBox()
        self.match_mode_combo.addItem("模糊匹配", 'fuzzy')
        self.match_mode_combo.addItem("前缀匹配", 'prefix')
        self.match_mode_combo.addItem("精确匹配", 'exact')
        self.match_mode_combo.currentIndexChanged.connect(self._schedule_search)
        row2.addWidget(QLabel("匹配:"))
        row2.addWidget(self.match_mode_combo, 1)

        self.sort_mode_combo = QComboBox()
        self.sort_mode_combo.addItem("文件名 A-Z", 'filename_asc')
        self.sort_mode_combo.addItem("文件名 Z-A", 'filename_desc')
        self.sort_mode_combo.addItem("子表数最多", 'sheet_count_desc')
        self.sort_mode_combo.addItem("子表数最少", 'sheet_count_asc')
        self.sort_mode_combo.currentIndexChanged.connect(self._on_sort_mode_changed)
        row2.addWidget(QLabel("排序:"))
        row2.addWidget(self.sort_mode_combo, 1)

        self.view_mode_combo = QComboBox()
        self.view_mode_combo.addItem("分组视图", 'grouped')
        self.view_mode_combo.addItem("列表视图", 'flat')
        self.view_mode_combo.currentIndexChanged.connect(self._on_view_mode_changed)
        row2.addWidget(QLabel("视图:"))
        row2.addWidget(self.view_mode_combo, 1)

        self.history_combo = QComboBox()
        self.history_combo.setMinimumWidth(220)
        self.history_combo.addItem("最近搜索")
        self.history_combo.currentIndexChanged.connect(self._on_history_selected)
        row2.addWidget(self.history_combo, 1)

        search_outer.addLayout(row2)

        self.index_status_label = QLabel("索引状态: 0 文件 / 0 子表 / 0 已深度索引")
        self.index_status_label.setStyleSheet("color: gray")
        search_outer.addWidget(self.index_status_label)

        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")

        # 结果区域（上：结果树 / 下：预览面板）
        self.splitter = QSplitter(Qt.Vertical)

        # 结果树
        self.result_tree = QTreeWidget()
        self.result_tree.setHeaderLabels(["文件名 / 子表", "命中子表数", "文件路径"])
        self.result_tree.setColumnWidth(0, 280)
        self.result_tree.setColumnWidth(1, 100)
        self.result_tree.setColumnWidth(2, 500)
        self.result_tree.setAlternatingRowColors(True)
        self.result_tree.setRootIsDecorated(True)

        # 绑定事件
        self.result_tree.itemClicked.connect(self._on_select)
        self.result_tree.itemDoubleClicked.connect(self._open_file)

        self.splitter.addWidget(self.result_tree)

        # 预览面板
        self.preview_container = QWidget()
        preview_layout = QVBoxLayout(self.preview_container)
        preview_layout.setContentsMargins(0, 4, 0, 0)
        preview_layout.setSpacing(2)

        self.preview_label = QLabel("预览: 请选择一个结果项")
        preview_layout.addWidget(self.preview_label)

        preview_search_layout = QHBoxLayout()
        preview_search_layout.setSpacing(6)
        self.preview_search_entry = QLineEdit()
        self.preview_search_entry.setPlaceholderText("当前预览内搜索")
        self.preview_search_entry.returnPressed.connect(self._search_within_preview)
        self.preview_search_entry.editingFinished.connect(self._search_within_preview)
        self.preview_search_entry.setEnabled(False)
        self.btn_preview_prev = QPushButton("上一个")
        self.btn_preview_prev.clicked.connect(self._goto_prev_preview_hit)
        self.btn_preview_prev.setEnabled(False)
        self.btn_preview_next = QPushButton("下一个")
        self.btn_preview_next.clicked.connect(self._goto_next_preview_hit)
        self.btn_preview_next.setEnabled(False)
        self.preview_hit_label = QLabel("命中: 0")
        self.preview_hit_label.setStyleSheet("color: gray")
        preview_search_layout.addWidget(QLabel("预览内搜索:"))
        preview_search_layout.addWidget(self.preview_search_entry, 1)
        preview_search_layout.addWidget(self.btn_preview_prev)
        preview_search_layout.addWidget(self.btn_preview_next)
        preview_search_layout.addWidget(self.preview_hit_label)
        preview_layout.addLayout(preview_search_layout)

        self.preview_table = QTableWidget()
        self.preview_table.setAlternatingRowColors(True)
        self.preview_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.preview_table.horizontalHeader().setStretchLastSection(False)
        self.preview_table.verticalHeader().setVisible(True)
        preview_layout.addWidget(self.preview_table)

        self.splitter.addWidget(self.preview_container)
        self.preview_visible = True
        self._saved_splitter_sizes = None

        # 预览面板折叠按钮
        self.btn_toggle_preview = QPushButton("▾ 折叠预览")
        self.btn_toggle_preview.setFixedWidth(120)
        self.btn_toggle_preview.clicked.connect(self._toggle_preview)
        preview_layout.insertWidget(0, self.btn_toggle_preview)

        self.splitter.setStretchFactor(0, 3)
        self.splitter.setStretchFactor(1, 2)

        main_layout.addWidget(self.splitter)

        # 底部操作按钮 + 右侧加载条
        bottom_widget = QWidget()
        bottom_layout = QHBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(10, 5, 10, 10)
        main_layout.addWidget(bottom_widget)

        self.btn_open = QPushButton("打开文件")
        self.btn_open.clicked.connect(self._open_file)
        self.btn_open.setEnabled(False)

        self.btn_locate = QPushButton("定位文件")
        self.btn_locate.clicked.connect(self._locate_file)
        self.btn_locate.setEnabled(False)

        self.btn_copy = QPushButton("复制路径")
        self.btn_copy.clicked.connect(self._copy_path)
        self.btn_copy.setEnabled(False)

        self.btn_export = QPushButton("导出结果")
        self.btn_export.clicked.connect(self._export_results)
        self.btn_export.setEnabled(False)

        bottom_layout.addWidget(self.btn_open)
        bottom_layout.addWidget(self.btn_locate)
        bottom_layout.addWidget(self.btn_copy)
        bottom_layout.addWidget(self.btn_export)
        bottom_layout.addStretch()

        self.scan_progress = QProgressBar()
        self.scan_progress.setFixedWidth(220)
        self.scan_progress.setRange(0, 100)
        self.scan_progress.setValue(0)
        self.scan_progress.setVisible(False)
        bottom_layout.addSpacing(12)
        bottom_layout.addWidget(self.scan_progress)

    def _toggle_preview(self):
        """折叠/展开预览面板"""
        if self.preview_visible:
            self._saved_splitter_sizes = self.splitter.sizes()
            self.preview_container.hide()
            self.preview_visible = False
            self.btn_toggle_preview.setText("▸ 展开预览")
            self.status_bar.showMessage("预览面板已折叠")
        else:
            self.preview_container.show()
            self.preview_visible = True
            if self._saved_splitter_sizes:
                self.splitter.setSizes(self._saved_splitter_sizes)
            self.btn_toggle_preview.setText("▾ 折叠预览")
            self.status_bar.showMessage("预览面板已展开", 2000)

    def keyPressEvent(self, event):
        """捕获 Ctrl+` / Cmd+` 折叠/展开预览"""
        key = event.key()
        mods = event.modifiers()
        is_ctrl = mods & Qt.CTRL
        is_cmd = mods & Qt.META
        # backtick: QuoteLeft (0x60), also check AsciiTilde on some layouts
        if (is_ctrl or is_cmd) and key in (Qt.Key_QuoteLeft, Qt.Key_AsciiTilde):
            self._toggle_preview()
            return
        super().keyPressEvent(event)

    def _check_existing_index(self):
        """检查是否存在已保存的扫描目录（可选显示）"""
        stats = self.index_manager.get_stats()
        if stats['file_count'] > 0:
            self._do_search()
        else:
            self._update_status_summary()

    def _restore_scan_directory(self):
        """恢复上次扫描目录"""
        saved_directory = self.settings.value('scan/last_directory', '')
        if saved_directory and os.path.isdir(saved_directory):
            self.scan_directory = saved_directory
            self.dir_label.setText(self._truncate_path(saved_directory))
            self.dir_label.setToolTip(saved_directory)
            return

        self.dir_label.setToolTip('')

    def _restore_ui_preferences(self):
        """恢复界面偏好设置"""
        self.match_mode_combo.blockSignals(True)
        self.sort_mode_combo.blockSignals(True)
        self.view_mode_combo.blockSignals(True)

        self._set_combo_by_data(
            self.match_mode_combo,
            self.settings.value('search/match_mode', 'fuzzy')
        )
        self._set_combo_by_data(
            self.sort_mode_combo,
            self.settings.value('search/sort_mode', self.current_sort_mode)
        )
        self._set_combo_by_data(
            self.view_mode_combo,
            self.settings.value('search/view_mode', self.current_view_mode)
        )

        self.match_mode_combo.blockSignals(False)
        self.sort_mode_combo.blockSignals(False)
        self.view_mode_combo.blockSignals(False)

        self.current_sort_mode = self.sort_mode_combo.currentData()
        self.current_view_mode = self.view_mode_combo.currentData()

    def _set_combo_by_data(self, combo_box, value):
        """按 data 值选中下拉项，避免启动时触发多余刷新"""
        for index in range(combo_box.count()):
            if combo_box.itemData(index) == value:
                combo_box.setCurrentIndex(index)
                return

    def _save_ui_preferences(self):
        """保存界面偏好设置"""
        self.settings.setValue('search/match_mode', self.match_mode_combo.currentData())
        self.settings.setValue('search/sort_mode', self.current_sort_mode)
        self.settings.setValue('search/view_mode', self.current_view_mode)

    def _refresh_index_status_label(self):
        """刷新索引覆盖率提示。"""
        status = self.index_manager.get_index_status()
        self.index_status_label.setText(
            "索引状态: "
            f"{status['file_count']} 文件 / "
            f"{status['sheet_count']} 子表 / "
            f"{status['indexed_cell_sheet_count']} 已深度索引 / "
            f"{status['pending_deep_index_count']} 待补全"
        )

    def _restore_action_buttons(self):
        has_selection = self.result_tree.currentItem() is not None
        for btn in [self.btn_open, self.btn_locate, self.btn_copy]:
            btn.setEnabled(has_selection)
        self.btn_export.setEnabled(bool(self.search_results))

    def _reset_preview_state(self):
        """清空当前预览和命中定位状态。"""
        self.preview_state = {
            'filepath': '',
            'sheet_name': '',
            'start_row': 1,
            'start_col': 1,
            'hits': [],
            'current_hit_index': -1,
            'active_keyword': '',
        }
        self._update_preview_controls()

    def _clear_preview(self, message="预览: 请选择一个结果项"):
        """重置预览表格内容。"""
        self.preview_label.setText(message)
        self.preview_table.clear()
        self.preview_table.setRowCount(0)
        self.preview_table.setColumnCount(0)
        self.preview_hit_label.setText("命中: 0")

    def _update_preview_controls(self):
        """根据当前命中状态刷新预览控件可用性和提示。"""
        has_sheet = bool(self.preview_state.get('filepath') and self.preview_state.get('sheet_name'))
        hits = self.preview_state.get('hits', [])
        current_index = self.preview_state.get('current_hit_index', -1)
        self.preview_search_entry.setEnabled(has_sheet)
        self.btn_preview_prev.setEnabled(len(hits) > 1)
        self.btn_preview_next.setEnabled(len(hits) > 1)

        if hits and current_index >= 0:
            self.preview_hit_label.setText(f"命中: {current_index + 1}/{len(hits)}")
        else:
            self.preview_hit_label.setText(f"命中: {len(hits)}")

    def _compute_preview_start(self, hit):
        """根据命中坐标计算预览窗口左上角。"""
        row = max(int(hit.get('row', 1)) - 3, 1)
        col = max(int(hit.get('col', 1)) - 2, 1)
        return row, col

    def _build_preview_label(self, filepath, sheet_name):
        """拼接预览标题和命中定位信息。"""
        base = f"预览: {sheet_name}  ({self._truncate_path(filepath, 80)})"
        hits = self.preview_state.get('hits', [])
        current_index = self.preview_state.get('current_hit_index', -1)
        if not hits:
            return base

        current_hit = hits[current_index] if 0 <= current_index < len(hits) else hits[0]
        from openpyxl.utils import get_column_letter
        coord = f"{get_column_letter(current_hit['col'])}{current_hit['row']}"
        return f"{base} | 命中 {len(hits)} 处 | 当前 {coord}"

    def _get_window_hit_positions(self, num_rows, num_cols):
        """返回当前预览窗口内所有命中的相对坐标。"""
        positions = set()
        start_row = self.preview_state.get('start_row', 1)
        start_col = self.preview_state.get('start_col', 1)
        end_row = start_row + num_rows - 1
        end_col = start_col + num_cols - 1

        for hit in self.preview_state.get('hits', []):
            row = hit.get('row', 0)
            col = hit.get('col', 0)
            if start_row <= row <= end_row and start_col <= col <= end_col:
                positions.add((row - start_row, col - start_col))

        return positions

    def _get_current_window_hit_position(self):
        """返回当前选中命中在预览窗口中的相对坐标。"""
        hits = self.preview_state.get('hits', [])
        current_index = self.preview_state.get('current_hit_index', -1)
        if not hits or current_index < 0 or current_index >= len(hits):
            return None

        hit = hits[current_index]
        start_row = self.preview_state.get('start_row', 1)
        start_col = self.preview_state.get('start_col', 1)
        return hit['row'] - start_row, hit['col'] - start_col

    def _save_scan_directory(self):
        """保存当前扫描目录"""
        if self.scan_directory:
            self.settings.setValue('scan/last_directory', self.scan_directory)

    def _import_sheet_aliases(self):
        """导入英文配置名到子表名的映射文件。"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            '选择映射文件',
            self.scan_directory or os.path.expanduser('~'),
            'Mapping Files (*.bat *.cmd *.txt);;All Files (*)'
        )
        if not file_path:
            return

        try:
            mappings = parse_sheet_alias_file(file_path)
            if not mappings:
                _msgbox(
                    self, 'warn', '提示',
                    '未在该文件中解析到有效映射。\n'
                    '标准格式: 英文配置名 子表名1 子表名2 ...\n'
                    '示例: TextConfig 界面文本 UI Text\n'
                    '也兼容旧格式: call do_conv.bat TextConfig 界面文本'
                )
                return

            imported_count = self.index_manager.replace_sheet_aliases(file_path, mappings)
            stats = self.index_manager.get_sheet_alias_stats()
            self.status_bar.showMessage(
                f"已导入映射：{imported_count} 条，当前共 {stats['alias_count']} 个英文名 / {stats['mapping_count']} 条映射"
            )
            _msgbox(
                self, 'info', '导入成功',
                f'已从以下文件导入映射:\n{file_path}\n\n'
                f'本次导入 {imported_count} 条映射。\n'
                f'现在可以直接用英文配置名搜索对应子表。'
            )
            self._do_search()
        except UnicodeDecodeError:
            _msgbox(self, 'error', '错误', '文件编码无法识别，请先保存为 UTF-8 编码后再导入')
        except Exception as e:
            _msgbox(self, 'error', '错误', f'导入映射失败: {e}')

    def _restore_search_history(self):
        """恢复最近搜索历史"""
        raw_history = self.settings.value('search/history', '[]')
        try:
            history = json.loads(raw_history)
        except (TypeError, json.JSONDecodeError):
            history = []

        self.search_history = [
            item for item in history
            if isinstance(item, dict)
            and (item.get('sheet_keyword') or item.get('filename_keyword') or item.get('cell_keyword'))
        ][:self.MAX_SEARCH_HISTORY]
        self._refresh_history_combo()

    def _save_search_history(self):
        """保存最近搜索历史"""
        self.settings.setValue('search/history', json.dumps(self.search_history, ensure_ascii=True))

    def _refresh_history_combo(self):
        """刷新最近搜索下拉框"""
        self.history_combo.blockSignals(True)
        self.history_combo.clear()
        self.history_combo.addItem("最近搜索")

        for item in self.search_history:
            self.history_combo.addItem(self._format_history_label(item), item)

        self.history_combo.setCurrentIndex(0)
        self.history_combo.blockSignals(False)

    def _format_history_label(self, item):
        """格式化搜索历史显示文本"""
        parts = []
        if item.get('sheet_keyword'):
            parts.append(f"子表:{item['sheet_keyword']}")
        if item.get('filename_keyword'):
            parts.append(f"文件:{item['filename_keyword']}")
        if item.get('cell_keyword'):
            parts.append(f"单元格:{item['cell_keyword']}")
        parts.append(self._match_mode_label(item.get('match_mode', 'fuzzy')))
        return ' | '.join(parts)

    def _match_mode_label(self, match_mode):
        """将匹配模式转成界面文案"""
        mapping = {
            'fuzzy': '模糊',
            'prefix': '前缀',
            'exact': '精确'
        }
        return mapping.get(match_mode, '模糊')

    def _record_search_history(self, sheet_keyword, filename_keyword, cell_keyword, match_mode):
        """记录最近搜索，避免输入过程中产生大量噪音"""
        if not sheet_keyword and not filename_keyword and not cell_keyword:
            return

        entry = {
            'sheet_keyword': sheet_keyword,
            'filename_keyword': filename_keyword,
            'cell_keyword': cell_keyword,
            'match_mode': match_mode
        }
        self.search_history = [item for item in self.search_history if item != entry]
        self.search_history.insert(0, entry)
        self.search_history = self.search_history[:self.MAX_SEARCH_HISTORY]
        self._save_search_history()
        self._refresh_history_combo()

    def _select_directory(self):
        """选择扫描目录"""
        directory = QFileDialog.getExistingDirectory(
            self, "选择要扫描的目录", os.path.expanduser("~")
        )
        if directory:
            self.scan_directory = directory
            self.dir_label.setText(self._truncate_path(directory))
            self.dir_label.setToolTip(directory)
            self._save_scan_directory()
            self._start_scan()

    def _rescan(self):
        """重新扫描"""
        if self.scan_directory:
            self._start_scan()
        else:
            _msgbox(self, 'info', '提示', '请先选择要扫描的目录')

    def _start_scan(self):
        """开始扫描（在线程中执行）"""
        if self.is_scanning:
            return

        self.is_scanning = True
        self.status_bar.showMessage("正在扫描...")
        self.scan_progress.setVisible(True)
        self.scan_progress.setRange(0, 0)

        # 禁用按钮
        for btn in [self.btn_open, self.btn_locate, self.btn_copy, self.btn_export]:
            btn.setEnabled(False)

        # 启动扫描线程
        self.scan_worker = ScanWorker(
            self.scan_directory, self.scanner, self.index_manager
        )
        self.scan_worker.finished.connect(self._on_scan_complete)
        self.scan_worker.error.connect(self._on_scan_error)
        self.scan_worker.progress.connect(self._on_scan_progress)
        self.scan_worker.start()

    def _on_scan_progress(self, current, total):
        """扫描进度回调"""
        if total > 0:
            self.scan_progress.setRange(0, total)
            self.scan_progress.setValue(current)
            self.status_bar.showMessage(f"正在扫描... {current}/{total}")
        else:
            self.scan_progress.setRange(0, 0)
            self.status_bar.showMessage("正在扫描...")

    def _on_scan_complete(self, added, updated, deleted, duration):
        """扫描完成回调"""
        self.is_scanning = False
        self.scan_progress.setVisible(False)
        self._restore_action_buttons()

        # 格式化耗时
        if duration >= 60:
            time_str = f"{duration / 60:.1f}分钟"
        elif duration >= 1:
            time_str = f"{duration:.1f}秒"
        else:
            time_str = f"{duration * 1000:.0f}毫秒"

        # 执行初始搜索
        self._do_search()
        self._refresh_index_status_label()
        self._pending_status_prefix = f"索引完成，耗时 {time_str}"

    def _on_scan_error(self, error_msg):
        """扫描/深度索引错误回调"""
        self.is_scanning = False
        self.scan_progress.setVisible(False)
        self._restore_action_buttons()
        self._refresh_index_status_label()
        self.status_bar.showMessage("操作出错")
        _msgbox(self, 'error', '错误', f'操作失败: {error_msg}')

    def _start_deep_index(self):
        """启动深度索引（提取单元格内容）"""
        if self.is_scanning:
            return
        self.is_scanning = True
        self.status_bar.showMessage("正在提取单元格内容...")
        self.scan_progress.setVisible(True)
        self.scan_progress.setRange(0, 0)

        for btn in [self.btn_open, self.btn_locate, self.btn_copy, self.btn_export]:
            btn.setEnabled(False)

        self.deep_worker = DeepIndexWorker(self.index_manager, self.scanner)
        self.deep_worker.finished.connect(self._on_deep_index_complete)
        self.deep_worker.error.connect(self._on_scan_error)
        self.deep_worker.progress.connect(self._on_scan_progress)
        self.deep_worker.start()

    def _on_deep_index_complete(self, processed, total, duration):
        """深度索引完成回调"""
        self.is_scanning = False
        self.scan_progress.setVisible(False)
        self._restore_action_buttons()

        # 格式化耗时（与扫描一致，通过 _pending_status_prefix 传递到最终状态栏）
        if duration >= 60:
            time_str = f"{duration / 60:.1f}分钟"
        elif duration >= 1:
            time_str = f"{duration:.1f}秒"
        else:
            time_str = f"{duration * 1000:.0f}毫秒"

        if total == 0:
            self._pending_status_prefix = "深度索引已是最新"
        else:
            self._pending_status_prefix = (
                f"深度索引完成：已处理 {processed}/{total} 个子表，耗时 {time_str}"
            )
            warning_summary = getattr(self.deep_worker, 'warning_summary', '')
            if warning_summary:
                first_line = warning_summary.splitlines()[0]
                self._pending_status_prefix += f"；{first_line}，详见日志"
        self._refresh_index_status_label()
        self._do_search()

    def _clear_index(self):
        """清空索引"""
        reply = _msgbox(self, 'question', '确认', '确定要清空所有索引数据吗？',
                        buttons=QMessageBox.Yes | QMessageBox.No,
                        default_button=QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.index_manager.clear_index()
            self.search_results = []
            self.result_tree.clear()
            self._reset_preview_state()
            self._refresh_index_status_label()
            self._update_status_summary(prefix='索引已清空')

    def _schedule_search(self):
        """防抖：延迟 75ms 后执行搜索，避免每次按键都触发查询"""
        self.search_timer.start(75)

    def _do_search(self):
        """在后台线程执行搜索，避免阻塞 UI"""
        # 取消上一次未完成的搜索（协作式：旧线程自行结束，不阻塞 UI）
        if self.search_worker and self.search_worker.isRunning():
            self.search_worker.finished.disconnect(self._on_search_finished)
            self.search_worker.error.disconnect(self._on_search_error)
            self.search_worker.cancel()

        sheet_keyword = self.sheet_entry.text().strip()
        filename_keyword = self.filename_entry.text().strip()
        cell_keyword = self.cell_entry.text().strip()
        match_mode = self.match_mode_combo.currentData()

        self.search_worker = SearchWorker(
            self.searcher, self.index_manager,
            sheet_keyword, filename_keyword, cell_keyword, match_mode
        )
        self.search_worker.finished.connect(self._on_search_finished)
        self.search_worker.error.connect(self._on_search_error)
        self.search_worker.start()

    def _on_search_finished(self, results):
        """后台搜索完成，在 UI 线程更新结果"""
        self.search_results = results
        self._sort_results()
        self._update_results()
        self._refresh_index_status_label()

    def _on_search_error(self, error_msg):
        """后台搜索出错"""
        _msgbox(self, 'error', '搜索错误', f'搜索失败: {error_msg}')

    def _on_search_committed(self):
        """仅在用户明确完成输入后写入搜索历史"""
        sheet_keyword = self.sheet_entry.text().strip()
        filename_keyword = self.filename_entry.text().strip()
        cell_keyword = self.cell_entry.text().strip()
        match_mode = self.match_mode_combo.currentData()
        self._record_search_history(sheet_keyword, filename_keyword, cell_keyword, match_mode)

    def _update_results(self, status_prefix=None):
        """更新搜索结果表格"""
        self.result_tree.clear()
        self._reset_preview_state()
        # 清空预览
        self._clear_preview(message="预览: 请选择一个结果项")

        for btn in [self.btn_open, self.btn_locate, self.btn_copy]:
            btn.setEnabled(False)
        self.btn_export.setEnabled(bool(self.search_results))

        if self.current_view_mode == 'flat':
            self._update_flat_results()
        else:
            self._update_grouped_results()

        effective_prefix = status_prefix if status_prefix is not None else self._pending_status_prefix
        self._pending_status_prefix = ''
        self._update_status_summary(prefix=effective_prefix)

    def _update_grouped_results(self):
        """按文件分组展示结果"""
        self.result_tree.setRootIsDecorated(True)
        self.result_tree.setHeaderLabels(["文件名 / 子表", "命中子表数", "文件路径"])
        self.result_tree.setUpdatesEnabled(False)

        for result in self.search_results:
            top_item = QTreeWidgetItem([
                result['filename'],
                str(result.get('sheet_count', 0)),
                result['filepath']
            ])
            top_item.setData(0, Qt.UserRole, result['filepath'])
            top_item.setData(0, Qt.UserRole + 1, 'file')
            self.result_tree.addTopLevelItem(top_item)

            for sheet_name in result.get('sheet_names', []):
                child_item = QTreeWidgetItem([
                    sheet_name,
                    '',
                    result['filepath']
                ])
                child_item.setData(0, Qt.UserRole, result['filepath'])
                child_item.setData(0, Qt.UserRole + 1, 'sheet')
                top_item.addChild(child_item)

            if result.get('sheet_names'):
                top_item.setExpanded(True)

        self.result_tree.setUpdatesEnabled(True)

    def _update_flat_results(self):
        """按旧版平铺列表展示结果"""
        self.result_tree.setRootIsDecorated(False)
        self.result_tree.setHeaderLabels(["文件名", "子表名称", "文件路径"])
        self.result_tree.setUpdatesEnabled(False)

        for result in self.search_results:
            sheet_names = result.get('sheet_names', [])
            if sheet_names:
                for sheet_name in sheet_names:
                    item = QTreeWidgetItem([
                        result['filename'],
                        sheet_name,
                        result['filepath']
                    ])
                    item.setData(0, Qt.UserRole, result['filepath'])
                    item.setData(0, Qt.UserRole + 1, 'flat')
                    self.result_tree.addTopLevelItem(item)
                continue

            item = QTreeWidgetItem([
                result['filename'],
                '',
                result['filepath']
            ])
            item.setData(0, Qt.UserRole, result['filepath'])
            item.setData(0, Qt.UserRole + 1, 'flat')
            self.result_tree.addTopLevelItem(item)

        self.result_tree.setUpdatesEnabled(True)

    def _on_select(self, item, column):
        """选中结果项"""
        self.btn_open.setEnabled(True)
        self.btn_locate.setEnabled(True)
        self.btn_copy.setEnabled(True)

        # 确定 filepath 和 sheet_name，触发预览
        filepath = item.data(0, Qt.UserRole)
        item_type = item.data(0, Qt.UserRole + 1)

        if item_type == 'sheet':
            sheet_name = item.text(0)
        elif item_type == 'file':
            # 顶级文件项：取该文件第一个匹配的子表
            for result in self.search_results:
                if result['filepath'] == filepath:
                    sheets = result.get('sheet_names', [])
                    sheet_name = sheets[0] if sheets else ''
                    break
            else:
                sheet_name = ''
        elif item_type == 'flat':
            sheet_name = item.text(1)
        else:
            sheet_name = ''

        if filepath and sheet_name:
            search_keyword = self.preview_search_entry.text().strip() or self.cell_entry.text().strip()
            match_mode = self.match_mode_combo.currentData()

            # 一次文件打开：命中查找 + 预览数据 + 表头
            hits, preview_data, header_row = self.scanner.read_sheet_with_hits(
                filepath, sheet_name,
                keyword=search_keyword or None,
                match_mode=match_mode,
            )

            current_hit_index = 0 if hits else -1
            if hits:
                start_row, start_col = self._compute_preview_start(hits[0])
            else:
                start_row, start_col = 2, 1

            self.preview_state.update({
                'filepath': filepath,
                'sheet_name': sheet_name,
                'start_row': start_row,
                'start_col': start_col,
                'hits': hits,
                'current_hit_index': current_hit_index,
                'active_keyword': search_keyword,
            })

            self.preview_search_entry.blockSignals(True)
            self.preview_search_entry.setText(search_keyword)
            self.preview_search_entry.blockSignals(False)
            self._update_preview_controls()
            data_start_row = start_row if start_row > 1 else 2
            self._render_preview_table(filepath, sheet_name, preview_data, header_row,
                                       data_start_row, start_col)
            return

        self._reset_preview_state()
        self._clear_preview(message="预览: 请选择一个子表")

    def _render_preview_table(self, filepath, sheet_name, data, header_row,
                               data_start_row, start_col):
        """将已加载的数据渲染到预览表格（不读文件）。"""
        self.preview_state['start_row'] = data_start_row
        self.preview_state['start_col'] = start_col
        self.preview_label.setText(self._build_preview_label(filepath, sheet_name))

        if not data:
            self._clear_preview(message=f"预览: {sheet_name} (空表)")
            self.status_bar.showMessage("就绪")
            return

        num_rows = len(data)
        num_cols = max((len(row) for row in data), default=0)

        self.preview_table.setRowCount(num_rows)
        self.preview_table.setColumnCount(num_cols)
        self.preview_table.clearContents()

        header = self.preview_table.horizontalHeader()
        for col in range(num_cols):
            header.setSectionResizeMode(col, QHeaderView.Interactive)
        header.setStretchLastSection(False)

        from openpyxl.utils import get_column_letter
        headers = []
        for i in range(num_cols):
            if i < len(header_row) and header_row[i].strip():
                headers.append(header_row[i])
            else:
                headers.append(get_column_letter(start_col + i))
        self.preview_table.setHorizontalHeaderLabels(headers)
        self.preview_table.setVerticalHeaderLabels(
            [str(data_start_row + i) for i in range(num_rows)]
        )

        window_hits = self._get_window_hit_positions(num_rows, num_cols)
        current_hit_position = self._get_current_window_hit_position()
        for r, row_data in enumerate(data):
            for c, cell_val in enumerate(row_data):
                item = QTableWidgetItem(cell_val)
                hit_key = (r, c)
                if hit_key in window_hits:
                    item.setBackground(QBrush(QColor('#FFF3B0')))
                if current_hit_position == hit_key:
                    item.setBackground(QBrush(QColor('#FFD166')))
                self.preview_table.setItem(r, c, item)

        self._apply_preview_column_widths(num_cols)
        self._update_preview_controls()
        self.preview_label.setText(self._build_preview_label(filepath, sheet_name))

        self.status_bar.showMessage("就绪")

    def _update_preview(self, filepath, sheet_name, start_row=1, start_col=1):
        """加载预览面板：读取 sheet 数据并渲染。用于命中跳转等已有 hits 的场景。"""
        if not filepath or not os.path.exists(filepath):
            self._clear_preview(message="预览: 文件不存在或已被移动")
            return

        data_start_row = start_row if start_row > 1 else 2
        self.status_bar.showMessage("正在加载预览...")
        QApplication.processEvents()
        try:
            _, preview_data, header_row = self.scanner.read_sheet_with_hits(
                filepath, sheet_name,
                keyword=None,
                start_row=data_start_row,
                start_col=start_col,
            )
        except Exception as e:
            self._clear_preview(message=f"预览: 读取失败 - {e}")
            self.status_bar.showMessage("预览加载失败")
            return

        self._render_preview_table(filepath, sheet_name, preview_data, header_row,
                                   data_start_row, start_col)

    def _search_within_preview(self):
        """在当前预览的 sheet 内重新搜索并跳到首个命中。"""
        filepath = self.preview_state.get('filepath')
        sheet_name = self.preview_state.get('sheet_name')
        if not filepath or not sheet_name:
            return

        keyword = self.preview_search_entry.text().strip()
        match_mode = self.match_mode_combo.currentData()

        # 一次文件打开：命中查找 + 预览数据 + 表头
        hits, preview_data, header_row = self.scanner.read_sheet_with_hits(
            filepath, sheet_name,
            keyword=keyword or None,
            match_mode=match_mode,
        )

        current_hit_index = 0 if hits else -1
        if hits:
            start_row, start_col = self._compute_preview_start(hits[0])
        else:
            start_row, start_col = 2, 1

        self.preview_state.update({
            'hits': hits,
            'current_hit_index': current_hit_index,
            'active_keyword': keyword,
            'start_row': start_row,
            'start_col': start_col,
        })
        data_start_row = start_row if start_row > 1 else 2
        self._render_preview_table(filepath, sheet_name, preview_data, header_row,
                                   data_start_row, start_col)

    def _goto_prev_preview_hit(self):
        """跳到上一个命中。"""
        self._goto_preview_hit(-1)

    def _goto_next_preview_hit(self):
        """跳到下一个命中。"""
        self._goto_preview_hit(1)

    def _goto_preview_hit(self, step):
        """在命中列表中跳转。"""
        hits = self.preview_state.get('hits', [])
        if not hits:
            return

        current_index = self.preview_state.get('current_hit_index', -1)
        if current_index < 0:
            current_index = 0
        else:
            current_index = (current_index + step) % len(hits)

        self.preview_state['current_hit_index'] = current_index
        start_row, start_col = self._compute_preview_start(hits[current_index])
        self._update_preview(
            self.preview_state['filepath'],
            self.preview_state['sheet_name'],
            start_row=start_row,
            start_col=start_col,
        )

    def _apply_preview_column_widths(self, num_cols):
        """按内容确定基础宽度，再把剩余空间尽量均分到所有列。"""
        if num_cols <= 0:
            return

        min_width = 80
        max_width = 240

        self.preview_table.resizeColumnsToContents()

        widths = []
        for col in range(num_cols):
            content_width = self.preview_table.columnWidth(col)
            widths.append(min(max(content_width, min_width), max_width))

        available_width = max(self.preview_table.viewport().width(), 0)
        total_width = sum(widths)

        if available_width > total_width:
            extra_width = available_width - total_width
            add_per_col, remainder = divmod(extra_width, num_cols)
            widths = [
                width + add_per_col + (1 if index < remainder else 0)
                for index, width in enumerate(widths)
            ]

        for col, width in enumerate(widths):
            self.preview_table.setColumnWidth(col, width)

    def _get_selected_filepath(self):
        """获取选中的文件路径"""
        selected_items = self.result_tree.selectedItems()
        if selected_items:
            return selected_items[0].data(0, Qt.UserRole)
        return None

    def _on_sort_mode_changed(self):
        """切换排序模式"""
        self.current_sort_mode = self.sort_mode_combo.currentData()
        self._save_ui_preferences()
        self._sort_results()
        self._update_results()

    def _on_view_mode_changed(self):
        """切换结果视图模式"""
        self.current_view_mode = self.view_mode_combo.currentData()
        self._save_ui_preferences()
        self._update_results()

    def _on_history_selected(self, index):
        """应用最近搜索历史"""
        if index <= 0:
            return

        item = self.history_combo.itemData(index)
        if not item:
            return

        self.sheet_entry.blockSignals(True)
        self.filename_entry.blockSignals(True)
        self.cell_entry.blockSignals(True)
        self.match_mode_combo.blockSignals(True)

        self.sheet_entry.setText(item.get('sheet_keyword', ''))
        self.filename_entry.setText(item.get('filename_keyword', ''))
        self.cell_entry.setText(item.get('cell_keyword', ''))
        self._set_combo_by_data(self.match_mode_combo, item.get('match_mode', 'fuzzy'))

        self.sheet_entry.blockSignals(False)
        self.filename_entry.blockSignals(False)
        self.cell_entry.blockSignals(False)
        self.match_mode_combo.blockSignals(False)

        self.history_combo.setCurrentIndex(0)
        self._do_search()

    def _sort_results(self):
        """按当前规则排序结果"""
        sort_mode = self.current_sort_mode or 'filename_asc'

        if sort_mode.startswith('sheet_count'):
            self.search_results.sort(
                key=lambda item: (item.get('sheet_count', 0), item['filename'].lower(), item['filepath'].lower())
            )
            if sort_mode == 'sheet_count_desc':
                self.search_results.reverse()
            return

        self.search_results.sort(
            key=lambda item: (item['filename'].lower(), item['filepath'].lower())
        )
        if sort_mode == 'filename_desc':
            self.search_results.reverse()

    def _update_status_summary(self, prefix: str = None):
        """更新结果统计"""
        file_count = len(self.search_results)
        matched_sheet_count = sum(result.get('sheet_count', 0) for result in self.search_results)
        if self.is_scanning:
            return

        index_status = self.index_manager.get_index_status()
        view_label = '分组视图' if self.current_view_mode == 'grouped' else '列表视图'
        summary = f"{view_label}：找到 {file_count} 个文件，{matched_sheet_count} 个子表命中"

        cell_keyword = self.cell_entry.text().strip() if hasattr(self, 'cell_entry') else ''
        sheet_keyword = self.sheet_entry.text().strip() if hasattr(self, 'sheet_entry') else ''
        filename_keyword = self.filename_entry.text().strip() if hasattr(self, 'filename_entry') else ''
        match_mode = self.match_mode_combo.currentData() if hasattr(self, 'match_mode_combo') else 'fuzzy'

        if index_status['file_count'] == 0:
            summary += " | 提示：请先选择目录并扫描建立索引"
        elif file_count == 0:
            if cell_keyword and index_status['pending_deep_index_count'] > 0:
                summary += (
                    " | 提示：未找到结果，且还有 "
                    f"{index_status['pending_deep_index_count']} 个子表未完成深度索引"
                )
            elif match_mode == 'exact' and (sheet_keyword or filename_keyword or cell_keyword):
                summary += " | 提示：当前为精确匹配，可尝试切换到模糊匹配"
            else:
                summary += (
                    " | 提示：当前索引中共有 "
                    f"{index_status['file_count']} 个文件 / {index_status['sheet_count']} 个子表，可尝试放宽关键词"
                )
        elif cell_keyword and index_status['pending_deep_index_count'] > 0:
            summary += (
                " | 提示：仍有 "
                f"{index_status['pending_deep_index_count']} 个子表未完成深度索引，单元格搜索可能不完整"
            )

        if prefix:
            self.status_bar.showMessage(f"{prefix}；{summary}")
            return
        self.status_bar.showMessage(summary)

    def _open_file(self):
        """打开文件"""
        filepath = self._get_selected_filepath()
        if filepath and os.path.exists(filepath):
            try:
                open_file(filepath)
            except Exception as e:
                _msgbox(self, 'error', '错误', f'无法打开文件: {e}')
        else:
            _msgbox(self, 'warn', '警告', '文件不存在或已被移动')

    def _locate_file(self):
        """在资源管理器中定位文件"""
        filepath = self._get_selected_filepath()
        if filepath and os.path.exists(filepath):
            try:
                open_in_explorer(filepath)
            except Exception as e:
                _msgbox(self, 'error', '错误', f'无法定位文件: {e}')
        else:
            _msgbox(self, 'warn', '警告', '文件不存在或已被移动')

    def _copy_path(self):
        """复制文件路径"""
        filepath = self._get_selected_filepath()
        if filepath:
            if copy_to_clipboard(filepath):
                _msgbox(self, 'info', '成功', '路径已复制到剪贴板')
            else:
                _msgbox(self, 'error', '错误', '复制失败')
        else:
            _msgbox(self, 'warn', '警告', '请先选择文件')

    def _export_results(self):
        """导出当前搜索结果为 CSV"""
        if not self.search_results:
            _msgbox(self, 'info', '提示', '当前没有可导出的搜索结果')
            return

        default_path = os.path.join(os.path.expanduser('~'), 'xlsx_search_results.csv')
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            '导出搜索结果',
            default_path,
            'CSV Files (*.csv)'
        )
        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['文件名', '子表名称', '文件路径', '命中子表数'])

                for result in self.search_results:
                    sheet_names = result.get('sheet_names', [])
                    if sheet_names:
                        for sheet_name in sheet_names:
                            writer.writerow([
                                result['filename'],
                                sheet_name,
                                result['filepath'],
                                result.get('sheet_count', 0)
                            ])
                        continue

                    writer.writerow([
                        result['filename'],
                        '',
                        result['filepath'],
                        result.get('sheet_count', 0)
                    ])

            self.status_bar.showMessage(f'已导出 {len(self.search_results)} 个文件到 {file_path}')
            _msgbox(self, 'info', '成功', f'搜索结果已导出到:\n{file_path}')
        except Exception as e:
            _msgbox(self, 'error', '错误', f'导出失败: {e}')

    def _truncate_path(self, path, max_length=50):
        """截断路径显示"""
        if len(path) <= max_length:
            return path
        return "..." + path[-max_length:]

    def closeEvent(self, event):
        """窗口关闭时保存偏好设置"""
        self._save_ui_preferences()
        self._save_scan_directory()
        event.accept()

    def eventFilter(self, obj, event):
        """顶部区域拖拽窗口 + 双击放大（macOS 统一标题栏）"""
        if sys.platform != 'darwin':
            return super().eventFilter(obj, event)

        if event.type() == event.MouseButtonDblClick:
            if self.isMaximized():
                self.showNormal()
            else:
                self.showMaximized()
            return True

        if event.type() == event.MouseButtonPress and event.button() == Qt.LeftButton:
            try:
                self.windowHandle().startSystemMove()
                return True
            except Exception:
                pass

        return super().eventFilter(obj, event)

    def run(self):
        """运行应用"""
        self.show()


def _get_icon_path():
    """获取图标文件路径，兼容开发环境和 PyInstaller 打包后的路径"""
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, 'icons', 'app_icon.png')


def _fix_macos_cfbundle_name():
    """在 QApplication 创建前修改 CFBundleName，防止 Qt 读取到 'Python'"""
    if sys.platform != 'darwin':
        return
    try:
        import ctypes
        import ctypes.util

        objc = ctypes.cdll.LoadLibrary(ctypes.util.find_library('objc'))
        objc.objc_getClass.restype = ctypes.c_void_p
        objc.objc_getClass.argtypes = [ctypes.c_char_p]
        objc.sel_registerName.restype = ctypes.c_void_p
        objc.sel_registerName.argtypes = [ctypes.c_char_p]

        def msg(restype, *argtypes):
            objc.objc_msgSend.restype = restype
            objc.objc_msgSend.argtypes = list(argtypes)
            return objc.objc_msgSend

        # 获取 NSBundle.mainBundle
        NSBundle = objc.objc_getClass(b'NSBundle')
        sel_main = objc.sel_registerName(b'mainBundle')
        bundle = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(NSBundle, sel_main)

        # 获取 infoDictionary
        sel_info = objc.sel_registerName(b'infoDictionary')
        info = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(bundle, sel_info)
        if not info:
            return

        # 造 NSString
        NSString = objc.objc_getClass(b'NSString')
        sel_str = objc.sel_registerName(b'stringWithUTF8String:')
        key = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_char_p)(
            NSString, sel_str, b'CFBundleName'
        )
        val = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_char_p)(
            NSString, sel_str, b'XlsxSearcher'
        )

        # 尝试通过 setValue:forKey: 直接修改 info dict（可能 immutable 则静默失败）
        sel_setValue = objc.sel_registerName(b'setValue:forKey:')
        msg(None, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(
            info, sel_setValue, val, key
        )
    except Exception:
        pass


def _apply_macos_unified_titlebar(window):
    """将 macOS 标题栏设置为透明统一样式（类似微信 macOS 端）"""
    if sys.platform != 'darwin':
        return
    try:
        import ctypes
        import ctypes.util

        objc = ctypes.cdll.LoadLibrary(ctypes.util.find_library('objc'))
        objc.objc_getClass.restype = ctypes.c_void_p
        objc.objc_getClass.argtypes = [ctypes.c_char_p]
        objc.sel_registerName.restype = ctypes.c_void_p
        objc.sel_registerName.argtypes = [ctypes.c_char_p]

        def msg(restype, *argtypes):
            objc.objc_msgSend.restype = restype
            objc.objc_msgSend.argtypes = list(argtypes)
            return objc.objc_msgSend

        # winId() -> NSView* -> [view window] -> NSWindow*
        view_ptr = int(window.winId())
        view = ctypes.c_void_p(view_ptr)
        sel_window = objc.sel_registerName(b'window')
        ns_window = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(view, sel_window)
        if not ns_window:
            return

        # 1. titlebarAppearsTransparent = YES
        sel_transparent = objc.sel_registerName(b'setTitlebarAppearsTransparent:')
        msg(None, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_byte)(ns_window, sel_transparent, True)

        # 2. styleMask |= NSWindowStyleMaskFullSizeContentView (1 << 15)
        sel_styleMask = objc.sel_registerName(b'styleMask')
        current_mask = msg(ctypes.c_ulong, ctypes.c_void_p, ctypes.c_void_p)(ns_window, sel_styleMask)
        NSWindowStyleMaskFullSizeContentView = 1 << 15
        sel_setStyleMask = objc.sel_registerName(b'setStyleMask:')
        msg(None, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_ulong)(
            ns_window, sel_setStyleMask, current_mask | NSWindowStyleMaskFullSizeContentView
        )

        # 3. movableByWindowBackground = YES (允许从内容区拖拽窗口)
        sel_movable = objc.sel_registerName(b'setMovableByWindowBackground:')
        msg(None, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_byte)(ns_window, sel_movable, True)

        # 4. 隐藏原生标题文字，避免与控件重叠
        # NSWindowTitleHidden = 1
        sel_setTitleVisibility = objc.sel_registerName(b'setTitleVisibility:')
        msg(None, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_long)(ns_window, sel_setTitleVisibility, 1)

    except Exception:
        pass


def _fix_macos_app_title(retry=0):
    """重命名 NSApplication 主菜单中的 App 菜单标题"""
    if sys.platform != 'darwin':
        return
    try:
        import ctypes
        import ctypes.util

        objc = ctypes.cdll.LoadLibrary(ctypes.util.find_library('objc'))
        objc.objc_getClass.restype = ctypes.c_void_p
        objc.objc_getClass.argtypes = [ctypes.c_char_p]
        objc.sel_registerName.restype = ctypes.c_void_p
        objc.sel_registerName.argtypes = [ctypes.c_char_p]

        def msg(restype, *argtypes):
            objc.objc_msgSend.restype = restype
            objc.objc_msgSend.argtypes = list(argtypes)
            return objc.objc_msgSend

        NSApp_cls = objc.objc_getClass(b'NSApplication')
        sel_shared = objc.sel_registerName(b'sharedApplication')
        shared = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(NSApp_cls, sel_shared)

        sel_mainMenu = objc.sel_registerName(b'mainMenu')
        main_menu = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(shared, sel_mainMenu)

        # 菜单可能还没创建好，重试
        if not main_menu:
            if retry < 10:
                QTimer.singleShot(100, lambda: _fix_macos_app_title(retry + 1))
            return

        sel_item = objc.sel_registerName(b'itemAtIndex:')
        app_item = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_long)(
            main_menu, sel_item, 0
        )
        if not app_item:
            return

        sel_submenu = objc.sel_registerName(b'submenu')
        app_menu = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(app_item, sel_submenu)
        if not app_menu:
            return

        NSString = objc.objc_getClass(b'NSString')
        sel_str = objc.sel_registerName(b'stringWithUTF8String:')
        title = msg(ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_char_p)(
            NSString, sel_str, b'XlsxSearcher'
        )

        sel_setTitle = objc.sel_registerName(b'setTitle:')
        msg(None, ctypes.c_void_p, ctypes.c_void_p, ctypes.c_void_p)(app_menu, sel_setTitle, title)

    except Exception:
        pass


def run_app():
    """启动应用程序"""
    _fix_macos_cfbundle_name()
    QApplication.setApplicationName("XlsxSearcher")
    app = QApplication(sys.argv)

    icon_path = _get_icon_path()
    if os.path.exists(icon_path):
        app.setWindowIcon(QIcon(icon_path))

    window = XlsxSearcherApp()

    if sys.platform == 'darwin':
        # 提前创建原生窗口句柄，在 show() 之前应用透明标题栏，避免闪烁
        _ = window.winId()
        _apply_macos_unified_titlebar(window)
        QTimer.singleShot(50, _fix_macos_app_title)

    window.show()
    sys.exit(app.exec_())
